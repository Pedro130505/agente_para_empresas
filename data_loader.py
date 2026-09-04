import openpyxl
import re
import unicodedata
from pathlib import Path
from config import EXCEL_PATH

def normalize_key(name):
    if not name:
        return ""
    name = str(name).strip().lower()
    # Remover acentos
    name = ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    # Erros comuns de digitação na planilha
    name = name.replace("arcelormitall", "arcelormittal")
    name = name.replace("sandivik", "sandvik")
    name = name.replace("grupo aterpa", "aterpa")
    name = name.replace("vinci energies", "vinci")
    name = name.replace("wabtec corporation", "wabtec")
    name = name.replace("mrv&co", "mrv")
    # Limpar caracteres não alfanuméricos exceto espaços
    name = re.sub(r'[^a-z0-9\s]', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def canonical_display_name(name):
    clean = str(name).strip()
    if clean.lower() in ["arcelormitall", "arcelormittal"]:
        return "ArcelorMittal"
    if clean.lower() in ["sandivik", "sandvik"]:
        return "Sandvik"
    if clean.lower() in ["grupo aterpa", "aterpa"]:
        return "Aterpa"
    if clean.lower() in ["vinci energies", "vinci"]:
        return "Vinci Energies"
    if clean.lower() in ["wabtec corporation", "wabtec"]:
        return "Wabtec"
    return clean

def load_companies_data(excel_path=EXCEL_PATH):
    if not Path(excel_path).exists():
        raise FileNotFoundError(f"Arquivo de dados não encontrado: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['2026']
    
    companies_map = {}
    
    # Helper para buscar ou criar entrada
    def get_or_create_entry(raw_name):
        canon_name = canonical_display_name(raw_name)
        key = normalize_key(canon_name)
        
        # Busca exata por chave normalizada
        if key in companies_map:
            return companies_map[key]
            
        # Nova entrada
        entry = {
            "nome": canon_name,
            "key": key,
            "participou_2024": "Não",
            "cota_2024": "N/A",
            "participou_2025": "Não",
            "cota_2025": "N/A",
            "participou_2026": "Não",
            "cota_2026": "N/A",
            "responsavel_2026": "",
            "nome_contato": "",
            "telefone": "",
            "email": "",
            "linkedin": ""
        }
        companies_map[key] = entry
        return entry

    # 1. Processar tabela de 2026 (Linhas 3 a 36)
    for r in range(3, 37):
        comp_name = ws.cell(r, 1).value
        status = ws.cell(r, 2).value
        cota_2026 = ws.cell(r, 3).value
        resp_2026 = ws.cell(r, 4).value
        
        if comp_name and str(comp_name).strip():
            entry = get_or_create_entry(comp_name)
            entry["participou_2026"] = "Sim" if cota_2026 else "Não"
            entry["cota_2026"] = str(cota_2026).strip() if cota_2026 else "N/A"
            if resp_2026:
                entry["responsavel_2026"] = str(resp_2026).strip()
            
    # 2. Processar tabela histórica 2024 / 2025 (Linhas 40 a 75)
    for r in range(40, ws.max_row + 1):
        comp_name = ws.cell(r, 2).value
        if not comp_name or not str(comp_name).strip():
            continue
            
        entry = get_or_create_entry(comp_name)
        
        contato = ws.cell(r, 3).value or ""
        telefone = ws.cell(r, 4).value or ""
        email = ws.cell(r, 5).value or ""
        linkedin = ws.cell(r, 6).value or ""
        
        p2024 = ws.cell(r, 7).value
        cota2024 = ws.cell(r, 8).value
        p2025 = ws.cell(r, 9).value
        cota2025 = ws.cell(r, 10).value
        
        part_2024 = "Sim" if str(p2024).strip().lower() in ["sim", "true", "1"] else "Não"
        part_2025 = "Sim" if str(p2025).strip().lower() in ["sim", "true", "1"] else "Não"
        
        str_cota2024 = str(cota2024).strip() if cota2024 and str(cota2024).strip() != "None" else "N/A"
        str_cota2025 = str(cota2025).strip() if cota2025 and str(cota2025).strip() != "None" else "N/A"
        
        entry["participou_2024"] = part_2024
        entry["cota_2024"] = str_cota2024 if part_2024 == "Sim" else "N/A"
        entry["participou_2025"] = part_2025
        entry["cota_2025"] = str_cota2025 if part_2025 == "Sim" else "N/A"
        
        if contato and not entry["nome_contato"]:
            entry["nome_contato"] = str(contato).strip()
        if telefone and not entry["telefone"]:
            entry["telefone"] = str(telefone).strip()
        if email and not entry["email"]:
            entry["email"] = str(email).strip()
        if linkedin and not entry["linkedin"]:
            entry["linkedin"] = str(linkedin).strip()

    return list(companies_map.values())

def find_or_create_company(company_name, companies_list=None):
    """Busca uma empresa na base do Excel ou cria um registro para nova prospecção."""
    if companies_list is None:
        companies_list = load_companies_data()
        
    target_key = normalize_key(company_name)
    
    # 1. Match exato
    for c in companies_list:
        if c["key"] == target_key:
            return c
            
    # 2. Match parcial
    for c in companies_list:
        if target_key in c["key"] or c["key"] in target_key:
            return c
            
    # 3. Empresa nova (fora do Excel da feira)
    canon_name = canonical_display_name(company_name)
    return {
        "nome": canon_name,
        "key": target_key,
        "origem": "Nova Prospecção",
        "participou_2024": "Não",
        "cota_2024": "N/A",
        "participou_2025": "Não",
        "cota_2025": "N/A",
        "participou_2026": "Não",
        "cota_2026": "N/A",
        "responsavel_2026": "",
        "status_2026": "",
        "nome_contato": "",
        "telefone": "",
        "email": "",
        "linkedin": ""
    }

if __name__ == "__main__":
    data = load_companies_data()
    print(f"Total de empresas consolidadas: {len(data)}\n")
    for item in data:
        print(f"{item['nome']:<25} | 2024: {item['participou_2024']} ({item['cota_2024']}) | 2025: {item['participou_2025']} ({item['cota_2025']}) | 2026: {item['participou_2026']} ({item['cota_2026']})")
